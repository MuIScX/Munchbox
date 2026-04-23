# backend/app/routes/import_data.py
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from app.db import get_db
from app.core.security import decode_token
from app.models import Menu, SaleData
from app.models.ingredient import Ingredient, IngredientHistory
import openpyxl
import datetime
import io
import csv
import pytz
bkk_tz = pytz.timezone("Asia/Bangkok")

router = APIRouter(prefix="/api/import", tags=["Import"])

# Map category name → menu type int (matching your existing type system)
CATEGORY_TYPE_MAP = {
    "CHICKEN": 1,
    "BURGERS": 1,
    "SNACKS": 2,
    "SIDE ITEMS": 2,
    "FRIED RICES": 1,
    "NON-ALCOHOL": 4,
    "ALCOHOL": 4,
    "SOUP": 2,
    "MAIN DISHES": 1,
    "SALAD": 2,
    "DESSERT & ICE CREAM": 3,
}

SKIP_ROWS = {"Sales by Item (Order Channel)", "Dine In,Take Out", "ITEM",
             "Composite Items", "Grand Total", "Total"}


def parse_sheet(ws, restaurant_id: int, db: Session):
    """Parse one sheet (one day) and return list of sale records created."""
    # Get date from row 2
    date_val = None
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        date_val = row[0]
        break
    if not isinstance(date_val, datetime.datetime):
        return 0

    sale_date = date_val.replace(hour=12, minute=0, second=0)
    current_category = None
    records_created = 0

    for row in ws.iter_rows(min_row=4, values_only=True):
        item_name, qty_raw, revenue_raw, _ = row

        if not item_name:
            continue

        # Category header row
        if qty_raw is None and item_name not in SKIP_ROWS:
            if not isinstance(item_name, datetime.datetime):
                current_category = item_name
            continue

        # Skip total/header rows
        if item_name in SKIP_ROWS:
            continue

        # Parse qty and revenue
        try:
            qty = int(str(qty_raw).replace(",", "")) if qty_raw else 0
            revenue = float(str(revenue_raw).replace(",", "")) if revenue_raw else 0.0
        except (ValueError, TypeError):
            continue

        if qty <= 0:
            continue

        # Calculate price per unit
        price = round(revenue / qty) if qty > 0 else 0

        # Get menu type from category
        menu_type = CATEGORY_TYPE_MAP.get(current_category, 1) if current_category else 1

        # Find or create menu item
        menu = db.query(Menu).filter(
            Menu.name == item_name,
            Menu.restaurant_id == restaurant_id,
        ).first()

        if not menu:
            menu = Menu(
                name=item_name,
                price=price,
                type=menu_type,
                restaurant_id=restaurant_id,
                is_active=1,
            )
            db.add(menu)
            db.flush()  # get menu.id without full commit
        else:
            # Update price if we have a better calculation
            if price > 0 and menu.price != price:
                menu.price = price

        # Check if sale already imported for this menu on this date (avoid duplicates)
        existing = db.query(SaleData).filter(
            SaleData.menu_id == menu.id,
            SaleData.restaurant_id == restaurant_id,
            SaleData.timestamp >= sale_date.replace(hour=0, minute=0, second=0),
            SaleData.timestamp <= sale_date.replace(hour=23, minute=59, second=59),
        ).first()

        if not existing:
            sale = SaleData(
                timestamp=sale_date,
                amount=qty,
                menu_id=menu.id,
                restaurant_id=restaurant_id,
            )
            db.add(sale)
            records_created += 1

    return records_created


@router.post("/sales")
def import_sales(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    identity: dict = Depends(decode_token),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are accepted")

    restaurant_id = identity["restaurantId"]

    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {str(e)}")

    total_records = 0
    sheets_processed = 0

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            records = parse_sheet(ws, restaurant_id, db)
            total_records += records
            sheets_processed += 1

        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

    return {
        "message": "success",
        "Data": {
            "sheets_processed": sheets_processed,
            "records_imported": total_records,
        }
    }


@router.post("/inventory")
def import_inventory(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    identity: dict = Depends(decode_token),
):
    """Import stock levels from CSV.
    Required columns: ingredient_name, new_stock
    Optional columns: as_of_date (YYYY-MM-DD), restock_type (before|after)
    """
    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only .csv files are accepted")

    restaurant_id = identity["restaurantId"]
    now = datetime.datetime.now(bkk_tz)

    try:
        contents = file.file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {str(e)}")

    required_cols = {"ingredient_name", "new_stock"}
    if reader.fieldnames is None or not required_cols.issubset(set(f.strip().lower() for f in reader.fieldnames)):
        raise HTTPException(status_code=400, detail=f"CSV must have columns: {', '.join(required_cols)}")

    updated = 0
    skipped = 0
    errors = []

    try:
        for i, row in enumerate(reader, start=2):
            name = (row.get("ingredient_name") or row.get("Ingredient Name") or "").strip()
            new_stock_raw = (row.get("new_stock") or row.get("New Stock") or "").strip()
            as_of_raw = (row.get("as_of_date") or row.get("As Of Date") or "").strip()
            restock_raw = (row.get("restock_type") or row.get("Restock Type") or "").strip().lower()

            if not name or not new_stock_raw:
                skipped += 1
                continue

            try:
                new_stock = float(new_stock_raw)
            except ValueError:
                errors.append(f"Row {i}: invalid new_stock '{new_stock_raw}'")
                continue

            ingredient = db.query(Ingredient).filter(
                Ingredient.restaurant_id == restaurant_id,
                Ingredient.name == name,
                Ingredient.is_active == 1,
            ).first()

            if not ingredient:
                errors.append(f"Row {i}: ingredient '{name}' not found")
                skipped += 1
                continue

            as_of_date = None
            if as_of_raw:
                try:
                    as_of_date = datetime.date.fromisoformat(as_of_raw)
                except ValueError:
                    pass

            restock_type = None
            if restock_raw == "before":
                restock_type = 1
            elif restock_raw == "after":
                restock_type = 2

            current = float(ingredient.stock_left)
            if new_stock < 0 or new_stock == current:
                skipped += 1
                continue

            action_type = 1 if new_stock > current else 2
            ingredient.stock_left = new_stock
            ingredient.last_update = now
            db.add(IngredientHistory(
                timestamp=now,
                action_type=action_type,
                amount=abs(new_stock - current),
                ingredient_id=ingredient.id,
                staff_id=0,
                restaurant_id=restaurant_id,
                new_current=int(new_stock),
                as_of_date=as_of_date,
                restock_type=restock_type,
            ))
            updated += 1

        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

    return {
        "message": "success",
        "Data": {"updated": updated, "skipped": skipped, "errors": errors[:10]},
    }